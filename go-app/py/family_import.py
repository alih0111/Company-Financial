# -*- coding: utf-8 -*-
"""
family_import.py — انتقال یک‌باره‌ی داده‌های اکسل «دارایی - 14050526.xlsx» به دیتابیس.

خوانده می‌شود:
  Sheet1: اشخاص، دارایی‌ها، تعداد/بهای تمام شده فعلی، قیمت‌ها (ستون C)، مانده حساب‌ها
  Sheet2: تاریخچه جمع کل (ستون‌های A/B/C/D) و آورده/برداشت‌ها (ستون‌های E/F/G)

اسکریپت idempotent است؛ اجرای مجدد، داده‌ها را با وضعیت اکسل جایگزین می‌کند
(تاریخچه MERGE می‌شود تا ردیف‌هایی که بعداً در سایت ثبت شده‌اند از بین نروند).
"""

import os
import sys
from pathlib import Path

import openpyxl
import pyodbc
from dotenv import load_dotenv

SCRIPT_DIR = Path(__file__).resolve().parent
load_dotenv(SCRIPT_DIR / ".env")
load_dotenv(SCRIPT_DIR.parent / ".env")
load_dotenv()

SERVER = os.getenv("DB_SERVER")
DATABASE = os.getenv("DB_NAME")
USERNAME = os.getenv("DB_USER")
PASSWORD = os.getenv("DB_PASSWORD")

EXCEL_PATH = SCRIPT_DIR.parent.parent / "دارایی - 14050526.xlsx"
DEFAULT_COMMISSION = 0.0088  # نرخ اکثر سهام در فرمول‌های اکسل

# نرخ کارمزد خاص چند دارایی (از فرمول‌های ستون ارزش اکسل استخراج شده)
ASSET_COMMISSIONS = {
    "مثقال": 0.001186,
    "یاقوت": 0.0001675,
}

# نماد بازار برای سینک خودکار قیمت از MarketPriceHistory.
# حامیا در BRS نیست و به‌صورت دستی قیمت می‌گیرد.
ASSET_SYMBOLS = {
    "مثقال": "مثقال",  # صندوق س.کالای آگاه
    "یاقوت": "یاقوت",  # صندوق س یاقوت آگاه-ثابت
    "کاسپین": "کاسپین",
    "زکشت": "زکشت",
    "دارو": "دارو",
    "کالا": "کالا",
    "سصوفی": "سصوفی",
    "دقاضی": "دقاضی",
    "کپرور": "کپرور",
    "شبهرن": "شبهرن",
    "گشان": "گشان",
}

# ────────────────────────────── ساخت جداول ──────────────────────────────

DDL = [
    """
    IF OBJECT_ID(N'dbo.FamilyPeople', N'U') IS NULL
    CREATE TABLE dbo.FamilyPeople (
        PersonID  INT IDENTITY(1,1) PRIMARY KEY,
        Name      NVARCHAR(100) NOT NULL,
        SortOrder INT NOT NULL DEFAULT 0,
        IsActive  BIT NOT NULL DEFAULT 1
    )""",
    """
    IF OBJECT_ID(N'dbo.FamilyAssets', N'U') IS NULL
    CREATE TABLE dbo.FamilyAssets (
        AssetID        INT IDENTITY(1,1) PRIMARY KEY,
        Name           NVARCHAR(100) NOT NULL,
        Symbol         NVARCHAR(50) NULL,
        Category       NVARCHAR(20) NOT NULL DEFAULT N'stock',
        CommissionRate FLOAT NOT NULL DEFAULT 0.0088,
        SortOrder      INT NOT NULL DEFAULT 0,
        IsActive       BIT NOT NULL DEFAULT 1
    )""",
    """
    IF OBJECT_ID(N'dbo.FamilyAssets', N'U') IS NOT NULL
       AND COL_LENGTH('dbo.FamilyAssets', 'CommissionRate') IS NULL
    ALTER TABLE dbo.FamilyAssets ADD CommissionRate FLOAT NOT NULL DEFAULT 0.0088""",
    """
    IF OBJECT_ID(N'dbo.FamilyAssets', N'U') IS NOT NULL
       AND COL_LENGTH('dbo.FamilyAssets', 'Symbol') IS NULL
    ALTER TABLE dbo.FamilyAssets ADD Symbol NVARCHAR(50) NULL""",
    """
    IF OBJECT_ID(N'dbo.FamilyHoldings', N'U') IS NULL
    CREATE TABLE dbo.FamilyHoldings (
        PersonID  INT NOT NULL,
        AssetID   INT NOT NULL,
        Quantity  FLOAT NOT NULL DEFAULT 0,
        CostBasis FLOAT NOT NULL DEFAULT 0,
        CONSTRAINT PK_FamilyHoldings PRIMARY KEY (PersonID, AssetID)
    )""",
    """
    IF OBJECT_ID(N'dbo.FamilyPrices', N'U') IS NULL
    CREATE TABLE dbo.FamilyPrices (
        DateKey NVARCHAR(10) NOT NULL,
        AssetID INT NOT NULL,
        Price   FLOAT NOT NULL,
        CONSTRAINT PK_FamilyPrices PRIMARY KEY (DateKey, AssetID)
    )""",
    """
    IF OBJECT_ID(N'dbo.FamilyAccounts', N'U') IS NULL
    CREATE TABLE dbo.FamilyAccounts (
        PersonID    INT PRIMARY KEY,
        CashBalance FLOAT NOT NULL DEFAULT 0
    )""",
    """
    IF OBJECT_ID(N'dbo.FamilyCashFlows', N'U') IS NULL
    CREATE TABLE dbo.FamilyCashFlows (
        ID        INT IDENTITY(1,1) PRIMARY KEY,
        DateKey   NVARCHAR(10) NOT NULL,
        Amount    FLOAT NOT NULL,
        Direction NVARCHAR(3) NOT NULL,
        Note      NVARCHAR(300) NULL
    )""",
    """
    IF OBJECT_ID(N'dbo.FamilyHistory', N'U') IS NULL
    CREATE TABLE dbo.FamilyHistory (
        DateKey     NVARCHAR(10) PRIMARY KEY,
        TotalValue  FLOAT NOT NULL,
        ChangeValue FLOAT NOT NULL DEFAULT 0,
        ChangePct   FLOAT NOT NULL DEFAULT 0,
        RecordedAt  DATETIME NOT NULL DEFAULT GETDATE()
    )""",
]


def get_connection():
    cs = (
        "DRIVER={ODBC Driver 17 for SQL Server};"
        f"SERVER={SERVER};"
        f"DATABASE={DATABASE};"
        f"UID={USERNAME};"
        f"PWD={PASSWORD};"
        "TrustServerCertificate=yes;"
    )
    return pyodbc.connect(cs)


# ────────────────────────────── خواندن اکسل ──────────────────────────────

# اشخاص و ستون‌های Sheet1: (نام، ستون تعداد، ستون بهای تمام شده)
PEOPLE_COLUMNS = [
    ("بابا", "D", "E"),
    ("مامان", "L", "M"),
    ("رضا", "T", "U"),
    ("وحیده", "AB", "AC"),
    ("فاطمه", "AJ", "AK"),
    ("علی", "AR", "AS"),
    ("وجیه", "BM", "BN"),
]

# مانده حساب‌ها: (نام شخص، سلول) — ردیف 18 برای اکثر اشخاص، BM19 برای وجیه
CASH_BALANCE_CELLS = [
    ("بابا", "F18"),
    ("مامان", "N18"),
    ("رضا", "V18"),
    ("وحیده", "AD18"),
    ("فاطمه", "AL18"),
    ("علی", "AT18"),
    ("وجیه", "BM19"),
]

ASSET_ROWS = range(3, 15)  # ردیف‌های 3 تا 14 در Sheet1
GOLD_NAMES = {"مثقال"}     # دسته طلا؛ بقیه سهام در نظر گرفته می‌شوند


def num(v):
    """تبدیل مقدار سلول به float؛ خطاهای #DIV/0! و متن‌ها صفر می‌شوند."""
    if v is None:
        return 0.0
    if isinstance(v, (int, float)):
        return float(v)
    return 0.0


def read_sheet1():
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    ws = wb["Sheet1"]

    assets = []  # (name, category, price)
    for row in ASSET_ROWS:
        name = (ws[f"B{row}"].value or "").strip()
        if not name:
            continue
        category = "gold" if name in GOLD_NAMES else "stock"
        assets.append((name, category, num(ws[f"C{row}"].value)))

    holdings = []  # (person, asset, qty, cost)
    for person, qty_col, cost_col in PEOPLE_COLUMNS:
        for row in ASSET_ROWS:
            name = (ws[f"B{row}"].value or "").strip()
            if not name:
                continue
            qty = num(ws[f"{qty_col}{row}"].value)
            cost = num(ws[f"{cost_col}{row}"].value)
            if qty != 0 or cost != 0:
                holdings.append((person, name, qty, cost))

    balances = {}  # person -> cash
    for person, cell in CASH_BALANCE_CELLS:
        balances[person] = num(ws[cell].value)

    return assets, holdings, balances


def read_sheet2():
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    ws = wb["Sheet2"]

    history = []   # (date, total, change, pct)
    cashflows = []  # (date, amount, direction, note)
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        date = row[0].value
        total = row[1].value
        if isinstance(date, str) and isinstance(total, (int, float)):
            date = date.strip()
            change = num(row[2].value) if row[2].value is not None else 0.0
            pct = num(row[3].value) if row[3].value is not None else 0.0
            history.append((date, float(total), change, pct))

        amount = row[4].value  # ستون E
        direction = row[5].value  # ستون F
        note = row[6].value  # ستون G
        if isinstance(date, str) and isinstance(amount, (int, float)) and direction in ("+", "-"):
            cashflows.append((date.strip(), float(amount),
                              "in" if direction == "+" else "out",
                              (note or "").strip() if isinstance(note, str) else ""))

    return history, cashflows


# ────────────────────────────── درج در دیتابیس ──────────────────────────────

def upsert_person(cur, name, sort_order):
    cur.execute(
        """
        MERGE dbo.FamilyPeople AS t
        USING (SELECT ? AS Name) AS s
        ON t.Name = s.Name
        WHEN MATCHED THEN UPDATE SET SortOrder = ?
        WHEN NOT MATCHED THEN INSERT (Name, SortOrder) VALUES (?, ?);
        """,
        name, sort_order, name, sort_order,
    )
    cur.execute("SELECT PersonID FROM dbo.FamilyPeople WHERE Name = ?", name)
    return cur.fetchone()[0]


def upsert_asset(cur, name, category, sort_order):
    commission = ASSET_COMMISSIONS.get(name, DEFAULT_COMMISSION)
    symbol = ASSET_SYMBOLS.get(name)
    cur.execute(
        """
        MERGE dbo.FamilyAssets AS t
        USING (SELECT ? AS Name) AS s
        ON t.Name = s.Name
        WHEN MATCHED THEN UPDATE
            SET Category = ?, CommissionRate = ?, SortOrder = ?, Symbol = ?
        WHEN NOT MATCHED THEN INSERT (Name, Category, CommissionRate, SortOrder, Symbol)
            VALUES (?, ?, ?, ?, ?);
        """,
        name, category, commission, sort_order, symbol,
        name, category, commission, sort_order, symbol,
    )
    cur.execute("SELECT AssetID FROM dbo.FamilyAssets WHERE Name = ?", name)
    return cur.fetchone()[0]


def compute_total_for_date(cur, date_key):
    """جمع کل با همان فرمول Go: دارایی‌ها × (1 − کارمزد دارایی) + مانده حساب‌ها."""
    cur.execute(
        """
        SELECT
            ISNULL(SUM(h.Quantity * pr.Price * (1 - a.CommissionRate)), 0),
            (SELECT ISNULL(SUM(CashBalance), 0) FROM dbo.FamilyAccounts)
        FROM dbo.FamilyHoldings h
        JOIN dbo.FamilyAssets a ON a.AssetID = h.AssetID
        OUTER APPLY (
            SELECT TOP 1 p.Price
            FROM dbo.FamilyPrices p
            WHERE p.AssetID = h.AssetID AND p.DateKey <= ? AND p.Price > 0
            ORDER BY p.DateKey DESC
        ) pr
        """,
        date_key,
    )
    holdings_value, total_cash = cur.fetchone()
    return holdings_value + total_cash


def main():
    if not EXCEL_PATH.exists():
        print(f"Excel file not found: {EXCEL_PATH}")
        sys.exit(1)

    print("Reading Excel ...")
    assets, holdings, balances = read_sheet1()
    history, cashflows = read_sheet2()
    print(f"  assets={len(assets)} holdings={len(holdings)} people={len(balances)}")
    print(f"  history rows={len(history)} cashflows={len(cashflows)}")

    last_date = history[-1][0] if history else None
    print(f"  last history date: {last_date}")

    conn = get_connection()
    conn.autocommit = False
    cur = conn.cursor()

    try:
        print("Creating tables (if missing) ...")
        for stmt in DDL:
            cur.execute(stmt)
        conn.commit()

        print("Upserting people & assets ...")
        person_ids, asset_ids = {}, {}
        for i, (person, _, _) in enumerate(PEOPLE_COLUMNS, start=1):
            person_ids[person] = upsert_person(cur, person, i)
        for i, (name, category, _) in enumerate(assets, start=1):
            asset_ids[name] = upsert_asset(cur, name, category, i)
        conn.commit()

        print("Replacing holdings & accounts ...")
        cur.execute("DELETE FROM dbo.FamilyHoldings")
        for person, asset, qty, cost in holdings:
            cur.execute(
                "INSERT INTO dbo.FamilyHoldings (PersonID, AssetID, Quantity, CostBasis) VALUES (?, ?, ?, ?)",
                person_ids[person], asset_ids[asset], qty, cost,
            )
        cur.execute("DELETE FROM dbo.FamilyAccounts")
        for person, _, _ in PEOPLE_COLUMNS:
            cur.execute(
                "INSERT INTO dbo.FamilyAccounts (PersonID, CashBalance) VALUES (?, ?)",
                person_ids[person], balances.get(person, 0.0),
            )
        conn.commit()

        if last_date:
            print(f"Saving prices for {last_date} ...")
            for name, _, price in assets:
                if price > 0:
                    cur.execute(
                        """
                        MERGE dbo.FamilyPrices AS t
                        USING (SELECT ? AS DateKey, ? AS AssetID) AS s
                        ON t.DateKey = s.DateKey AND t.AssetID = s.AssetID
                        WHEN MATCHED THEN UPDATE SET Price = ?
                        WHEN NOT MATCHED THEN INSERT (DateKey, AssetID, Price) VALUES (?, ?, ?);
                        """,
                        last_date, asset_ids[name], price, last_date, asset_ids[name], price,
                    )
            conn.commit()

        print("Merging history ...")
        for date, total, change, pct in history:
            cur.execute(
                """
                MERGE dbo.FamilyHistory AS t
                USING (SELECT ? AS DateKey) AS s
                ON t.DateKey = s.DateKey
                WHEN NOT MATCHED THEN INSERT (DateKey, TotalValue, ChangeValue, ChangePct)
                    VALUES (?, ?, ?, ?);
                """,
                date, date, total, change, pct,
            )
        conn.commit()

        print("Replacing cash flows ...")
        cur.execute("DELETE FROM dbo.FamilyCashFlows")
        for date, amount, direction, note in cashflows:
            cur.execute(
                "INSERT INTO dbo.FamilyCashFlows (DateKey, Amount, Direction, Note) VALUES (?, ?, ?, ?)",
                date, amount, direction, note,
            )
        conn.commit()

        # جمع آخرین روز با فرمول جدید (شامل وجیه) دوباره محاسبه می‌شود تا
        # ادامه‌ی تاریخچه در سایت با همین مبنا محاسبه شود.
        if last_date:
            total = compute_total_for_date(cur, last_date)
            cur.execute(
                "UPDATE dbo.FamilyHistory SET TotalValue = ? WHERE DateKey = ?",
                total, last_date,
            )
            conn.commit()
            print(f"Recomputed {last_date} total with site formula: {total:,.0f}")

        print("Done.")

    except Exception:
        conn.rollback()
        raise
    finally:
        conn.close()


if __name__ == "__main__":
    main()
